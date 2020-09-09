def fix_gbk(object):
    """
    修复gbk编码问题
    'gbk' codec can't encode character '\xa0' in position 202: illegal multibyte sequence
    """
    return object.replace('\xa0', '') if isinstance(object, str) else object


def trans_excel_to_csv(excel_path, csv_path, encoding="utf8", gbk_fixing=True):
    import csv
    import openpyxl
    with open(csv_path, mode="w", newline='', encoding=encoding) as fw:
        writer = csv.writer(fw, delimiter=',')
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for row in ws.rows:
            if gbk_fixing:
                writer.writerow([fix_gbk(c.value) for c in row])
            else:
                writer.writerow([c.value for c in row])


def read_csv_dict(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else ["utf8", "utf-8-sig", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg=f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding=e) as fr:
                reader = csv.reader(fr)
                head = next(reader)
                for row in reader:
                    yield dict(zip(head, row))
            return
        except:
            pass


def read_csv_head(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else ["utf8", "utf-8-sig", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg=f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding = e) as fr:
                reader=csv.reader(fr)
                return next(reader)
        except:
            pass


def read_csv_rows(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else ["utf8", "utf-8-sig", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg=f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding = e) as fr:
                reader=csv.reader(fr)
                next(reader)
                yield from reader
            return
        except:
            pass


def open_excel(path):
    import xlrd
    return xlrd.open_workbook(path)


def get_excel_sheet(excel, sheet_name = None):
    sheet_names=excel.sheet_names()
    if sheet_name is None:
        sheet_name=sheet_names[0]
    assert sheet_name in sheet_names
    return excel.sheet_by_name(sheet_name)


def read_excel_head(excel, sheet_name = None):
    st=get_excel_sheet(excel, sheet_name = sheet_name)
    return st.row_values(0)


def read_excel_rows(excel, sheet_name = None):
    st=get_excel_sheet(excel, sheet_name = sheet_name)
    nrows=st.nrows
    yield from (st.row_values(i) for i in range(1, nrows))


def read_excel_items(excel, sheet_name = None):
    st=get_excel_sheet(excel, sheet_name = sheet_name)
    head=read_excel_head(excel, sheet_name = sheet_name)
    nrows=st.nrows
    yield from (dict(zip(head, st.row_values(i))) for i in range(1, nrows))

# def save_data_to_excel(data, excel_path, skip_error_row=False):
#     import openpyxl
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     for row in data:
#         try:
#             ws.append(row)
#         except Exception as ex:
#             if not skip_error_row:
#                 raise ex
#     wb.save(excel_path)


# def read_xls(filepath, sheetname=None):
#     import xlrd
#     from datetime import datetime
#     rbook = xlrd.open_workbook(filepath)
#     if sheetname:
#         sheet = rbook.sheet_by_name(sheetname)
#     else:
#         sheet = rbook.sheets()[0]
#     rows = sheet.nrows
#     cols = sheet.ncols
#     all_content = []
#     for i in range(rows):
#         row_content = []
#         for j in range(cols):
#             ctype = sheet.cell(i, j).ctype  # 表格的数据类型
#             cell = sheet.cell_value(i, j)
#             if ctype == 2 and cell % 1 == 0:  # 如果是整形
#                 cell = int(cell)
#             elif ctype == 3:
#                 # 转成datetime对象
#                 date = datetime(*xlrd.xldate_as_tuple(cell, 0))
#                 cell = date.strftime('%Y-%m-%d %H:%M:%S')
#             elif ctype == 4:
#                 cell = True if cell == 1 else False
#             row_content.append(cell)
#         all_content.append(row_content)
#         # print('[' + ','.join("'" + str(element) + "'" for element in row_content) + ']')
#     return all_content


# def get_data_from_excel(excel_path):
#     if excel_path.endswith(".xlsx"):
#         import openpyxl
#         wb = openpyxl.load_workbook(excel_path)
#         ws = wb.active
#         return list(ws.values)
#     else:
#         return read_xls(excel_path)


# def get_data_from_excel_streamingly(filepath):
#     if filepath.endswith(".xlsx"):
#         import openpyxl
#         wb = openpyxl.load_workbook(filepath)
#         ws = wb.active
#         for r in ws.values:
#             yield r
#     else:
#         import xlrd
#         from datetime import datetime
#         rbook = xlrd.open_workbook(filepath)
#         sheet = rbook.sheets()[0]
#         rows = sheet.nrows
#         cols = sheet.ncols
#         for i in range(rows):
#             row_content = []
#             for j in range(cols):
#                 ctype = sheet.cell(i, j).ctype  # 表格的数据类型
#                 cell = sheet.cell_value(i, j)
#                 if ctype == 2 and cell % 1 == 0:  # 如果是整形
#                     cell = int(cell)
#                 elif ctype == 3:
#                     # 转成datetime对象
#                     date = datetime(*xlrd.xldate_as_tuple(cell, 0))
#                     cell = date.strftime('%Y-%m-%d %H:%M:%S')
#                 elif ctype == 4:
#                     cell = True if cell == 1 else False
#                 row_content.append(cell)
#             yield row_content
