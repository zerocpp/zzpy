DEFAULT_ROOT_DIR = "output"


def remove_dir(dir_path):
    """删除文件夹"""
    import os
    import shutil
    if os.path.exists(dir_path):
        try:
            os.removedirs(dir_path)
        except Exception as ex:
            shutil.rmtree(dir_path, ignore_errors=True)


def create_dir(dir_path):
    """创建文件夹"""
    import os
    import shutil
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def init_dir(dir_path):
    """初始化（先删除再创建）文件夹"""
    remove_dir(dir_path)
    create_dir(dir_path)


def get_one_file_path_from_dir(dir_path):
    """返回文件夹中（第一个）文件的地址"""
    import os
    for parent, dirnames, filenames in os.walk(dir_path):
        for filename in filenames:
            return os.path.join(parent, filename)
    return None


def move_file_from_src_dir(dst_file_path, src_dir_path):
    """从src_dir中拿到（唯一）文件，移动到dst_dir，并重命名"""
    import shutil
    import os
    src_file_path = get_one_file_path_from_dir(src_dir_path)
    if not src_file_path or not os.path.exists(src_file_path):
        return
    if not os.path.exists(os.path.dirname(dst_file_path)):
        create_dir(os.path.dirname(dst_file_path))
    shutil.move(src_file_path, dst_file_path)


def get_file_line_count(file_path):
    """获取文件行数"""
    count = 0
    with open(file_path, 'rb') as fr:
        for _ in enumerate(fr):
            count += 1
    return count


def get_file_name_list_from_dir(dir_path):
    """获取文件夹下所有文件名"""
    import os
    file_names = []
    for parent_dir, dirs, files in os.walk(dir_path):
        for file_name in files:
            file_names.append(file_name)
    return file_names


def get_file_path_list_from_dir(dir_path):
    """获取文件夹下所有文件路径"""
    import os
    file_pathes = []
    for parent_dir, dirs, files in os.walk(dir_path):
        for file_name in files:
            file_path = os.path.join(parent_dir, file_name)
            file_pathes.append(file_path)
    return file_pathes


def _read_file_using_encoding(file_path, encoding):
    """按编码读取文件"""
    f = None
    text = None
    try:
        f = open(file_path, encoding=encoding)
        text = f.read()
    except:
        pass
    finally:
        if f:
            f.close()
    return text


def read_bin_file(file_path):
    """读取二进制文件"""
    data = None
    with open(file_path, mode="rb") as fr:
        data = fr.read()
    return data


def read_file(file_path):
    """读取文件"""
    for encoding in ("utf8", "gbk"):
        text = _read_file_using_encoding(file_path, encoding=encoding)
        if text is not None:
            return text
    return None


def write_file(content, file_path, encoding=None):
    """写文件"""
    import os
    if encoding is None:
        encoding = "utf8"
    try:
        dir_path = os.path.dirname(file_path)
        if not os.path.exists(dir_path):
            create_dir(dir_path)
    except:
        pass
    with open(file_path, mode="w", encoding=encoding) as fw:
        fw.write(content)


def get_tmp_dir_path(root_dir=None):
    """临时目录"""
    if root_dir is None:
        root_dir = DEFAULT_ROOT_DIR
    import os
    return os.path.join(root_dir, "tmp")


def get_work_dir_path(root_dir=None):
    """工作目录"""
    if root_dir is None:
        root_dir = DEFAULT_ROOT_DIR
    import os
    return os.path.join(root_dir, "work")


def init_dirs(root_dir=None):
    '''初始化所有目录：临时目录、工作目录'''
    if root_dir is None:
        root_dir = DEFAULT_ROOT_DIR
    for dir_path_func in (get_tmp_dir_path, get_work_dir_path):
        dir_path = dir_path_func(root_dir=root_dir)
        init_dir(dir_path)


def read_jsonline(file_path):
    """从jsonline文件按行读取"""
    import jsonlines
    with open(file_path, encoding='utf8') as fr:
        yield from jsonlines.Reader(fr)


def read_jsonline_with_progressbar(file_path, title=None):
    """从jsonline文件按行读取，带进度条"""
    import jsonlines
    from .zprogress import pb
    with open(file_path, encoding='utf8') as fr:
        yield from pb(jsonlines.Reader(fr), total=get_file_line_count(file_path), title=title if title else "进度")


def _split_path_list(path_list, split_pattern):
    component_list = []
    for path in path_list:
        component_list.extend(path.split(split_pattern))
    return component_list


def normalize_path(path):
    """标准化路径"""
    import os
    components = [path]
    for split_pattern in ("/", "\\"):
        components = _split_path_list(components, split_pattern=split_pattern)
    prefix = "/" if path.startswith("/") else ""
    return prefix + os.path.join(*list(components))


def download_file(url, path):
    import requests
    resp = requests.get(url)
    with open(path, "wb") as fw:
        fw.write(resp.content)


# excel


def remove_illegal_characters(content):
    import re
    ILLEGAL_CHARACTERS_RE = re.compile(
        r'[\000-\010]|[\013-\014]|[\016-\037]|\xa0')
    content = ILLEGAL_CHARACTERS_RE.sub(r'', content)
    return content


def trans_excel_to_csv(excel_path, csv_path, encoding="utf8", gbk_fixing=True):
    import csv
    import openpyxl
    with open(csv_path, mode="w", newline='', encoding=encoding) as fw:
        writer = csv.writer(fw, delimiter=',')
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for row in ws.rows:
            if gbk_fixing:
                writer.writerow([remove_illegal_characters(c.value)
                                 for c in row])
            else:
                writer.writerow([c.value for c in row])


def trans_csv_to_excel(csv_path, excel_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(read_csv_head(csv_path))
    for row in read_csv_rows(csv_path):
        ws.append(row)
    wb.save(excel_path)


def read_csv_dict(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else [
                 "utf-8-sig", "utf8", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg = f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding=e) as fr:
                reader = csv.reader(fr)
                head = next(reader)
                for row in reader:
                    yield dict(zip(head, row))
            return
        except:
            pass


def read_csv_head(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else [
                 "utf-8-sig", "utf8", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg = f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding=e) as fr:
                reader = csv.reader(fr)
                return next(reader)
        except:
            pass


def read_csv_rows(path, encoding=None):
    import csv
    unknown_encoding = "unknown"
    encodings = ([encoding] if encoding else [
                 "utf-8-sig", "utf8", "gbk"]) + [unknown_encoding]
    for e in encodings:
        if e == unknown_encoding:
            error_msg = f"文件编码错误: {path}"
            raise Exception(error_msg)
        try:
            with open(path, encoding=e) as fr:
                reader = csv.reader(fr)
                next(reader)
                yield from reader
            return
        except:
            pass


def open_excel(path):
    import xlrd
    return xlrd.open_workbook(path)


def save_excel_rows(rows, path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def save_excel_items(items, path, head=None):
    import openpyxl
    items = list(items)
    if not head:
        head = list(items[0].keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(head)
    for item in items:
        row = [item.get(k, "") for k in head]
        ws.append(row)
    wb.save(path)


def get_excel_sheet(excel, sheet_name=None):
    sheet_names = excel.sheet_names()
    if sheet_name is None:
        sheet_name = sheet_names[0]
    assert sheet_name in sheet_names
    return excel.sheet_by_name(sheet_name)


def read_excel_head(excel, sheet_name=None):
    st = get_excel_sheet(excel, sheet_name=sheet_name)
    return st.row_values(0)


def read_excel_rows(excel, sheet_name=None):
    st = get_excel_sheet(excel, sheet_name=sheet_name)
    nrows = st.nrows
    yield from (st.row_values(i) for i in range(1, nrows))


def read_excel_items(excel, sheet_name=None):
    st = get_excel_sheet(excel, sheet_name=sheet_name)
    head = read_excel_head(excel, sheet_name=sheet_name)
    nrows = st.nrows
    yield from (dict(zip(head, st.row_values(i))) for i in range(1, nrows))


def save_items_to_csv(items, path, head=None):
    import csv
    with open(path, mode="w", encoding="utf-8", newline="") as fw:
        writer = csv.writer(fw)
        if head is None:
            if len(items) > 0:
                head = list(items[0].keys())
        if head:
            writer.writerow(head)
        for it in items:
            writer.writerow([it.get(k, "") for k in head])


def stringify(object):
    return str(object) if object else ""


def convert_csv_to_xlsx(csv_path, xlsx_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(read_csv_head(csv_path))
    for row in read_csv_rows(csv_path):
        ws.append([remove_illegal_characters(
            stringify(i)) if i else "" for i in row])
    wb.save(xlsx_path)


def convert_xlsx_to_csv(xlsx_path, csv_path, encoding="utf-8"):
    import csv
    import openpyxl
    with open(csv_path, mode="w", newline='', encoding=encoding) as fw:
        writer = csv.writer(fw, delimiter=',')
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for row in ws.rows:
            writer.writerow([remove_illegal_characters(
                stringify(c.value)) for c in row])


def convert_xlsx_to_jsonl(xlsx_path, jsonl_path, force_str=True):
    import jsonlines
    with jsonlines.open(jsonl_path, mode="w") as fw:
        items = read_excel_items(open_excel(xlsx_path))
        for it in items:
            if force_str:
                for k in it:
                    it[k] = str(it[k])
            fw.write(it)


def convert_jsonl_to_xlsx(jsonl_path, xlsx_path):
    save_excel_items(items=list(read_jsonline(jsonl_path)), path=xlsx_path)


def save_items_to_jsonl(items, path):
    import jsonlines
    with jsonlines.open(path, mode="w") as fw:
        for it in items:
            fw.write(it)
